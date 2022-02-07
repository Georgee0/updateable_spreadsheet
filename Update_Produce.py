# Update produce.py

import openpyxl

wb = openpyxl.load_workbook('C:/Users/Georgee/Documents/automate_online-materials/produceSales.xlsx')
sheet = wb['Sheet']

# Produce price and their updated prices
PRICE_UPDATE = {'Garlic': 3.07,
                    'Celery':1.19,
                    'Lemon': 1.27}

# Loop through the row and update the prices 
for rowNum in range(2,sheet.max_row):  # Skip the first row
    ProduceName = sheet.cell(row=rowNum, column=1).value
    if ProduceName in PRICE_UPDATE:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATE[ProduceName]

wb.save('C:/Users/Georgee/Documents/Python VS_code/UpdatedSalesList.xlsx')        
